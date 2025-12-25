"""知识库（RAG）相关服务。

目标：
- 仅支持 Excel（.xlsx/.xlsm）入库（默认只处理第一个 sheet）
- 默认使用第一列作为主标签（primary_tag），支持严格等值命中增强检索
- 真实接入 Chroma：chunk（行）级向量存储、覆盖时删除旧 chunk，避免残留
- 检索-only：不接 LLM，返回证据列表 + 规则化摘要
- 缓存：优先 Redis，不可用时退化到内存 TTL；以 collection revision 自动失效
"""
from __future__ import annotations

import json
import logging
import time
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook

from config import get_settings
from database import session_scope
from models import KnowledgeCollection, KnowledgeDocument, RAGQuery, RagDocumentLog, LogLevel
from services.rag_cache import HybridCache
from services.rag_normalize import extract_candidate_primary_tag, extract_candidate_primary_tag_kv, normalize_primary_tag_value, normalize_text
from services.rag_vector_store import ChromaVectorStore, VectorHit, embed_query, embed_texts_batched

logger = logging.getLogger(__name__)


class DocumentAlreadyExistsError(RuntimeError):
    pass


@dataclass(frozen=True)
class ExcelIndexReport:
    primary_tag_key: str
    sheet_name: str
    rows_total: int
    rows_indexed: int
    rows_skipped_empty_tag: int
    primary_tag_unique: int


class RagService:
    """提供知识库增删查等操作。"""

    def __init__(self) -> None:
        self._settings = get_settings()
        self._cache = HybridCache(self._settings.redis_url)
        self._vector = ChromaVectorStore()

    def create_collection(self, name: str, description: str | None, user_id: int) -> KnowledgeCollection:
        with session_scope() as session:
            collection = KnowledgeCollection(
                name=name,
                description=description or "",
                document_count=0,
                collection_type="document",
                is_active=True,
                created_by=user_id,
            )
            session.add(collection)
            session.flush()
            # 绑定稳定的 chroma_collection_id（避免名称变更导致索引丢失）
            if collection.id is None:
                # 正常情况下 flush 后一定有主键；若没有，说明 DB 底层异常
                raise RuntimeError("知识库创建失败：未分配主键")
            chroma_id = collection.chroma_collection_id or f"kc_{collection.id}"
            collection.chroma_collection_id = chroma_id
            collection.embedding_model = collection.embedding_model or "bge-base-zh-v1.5"
            session.add(collection)
            session.refresh(collection)
            # 确保 Chroma collection 存在
            self._vector.get_or_create_collection(chroma_id)
            return collection

    def get_collection(self, collection_id: int) -> KnowledgeCollection | None:
        with session_scope() as session:
            return session.get(KnowledgeCollection, collection_id)

    def list_collections(self, user_id: int | None = None) -> Sequence[KnowledgeCollection]:
        with session_scope() as session:
            query = session.query(KnowledgeCollection).order_by(KnowledgeCollection.created_at.desc())
            if user_id is not None:
                query = query.filter(
                    (KnowledgeCollection.created_by == user_id) | (KnowledgeCollection.created_by.is_(None))
                )
            return query.all()

    def get_document(self, document_id: int) -> KnowledgeDocument | None:
        with session_scope() as session:
            return session.get(KnowledgeDocument, document_id)

    def list_document_logs(self, document_id: int, limit: int = 200) -> Sequence[RagDocumentLog]:
        with session_scope() as session:
            return (
                session.query(RagDocumentLog)
                .filter(RagDocumentLog.document_id == document_id)
                .order_by(RagDocumentLog.created_at.asc())
                .limit(limit)
                .all()
            )

    def _append_document_log(self, document_id: int, level: LogLevel, message: str, meta: dict | None = None) -> None:
        # 同步输出到后端日志（便于在服务端排查，不依赖前端）
        try:
            meta_str = json.dumps(meta, ensure_ascii=False) if meta else ""
        except Exception:
            meta_str = ""
        lvl = str(level.value if hasattr(level, "value") else level).upper()
        if lvl == "DEBUG":
            logger.debug("[RAG][doc=%s] %s %s", document_id, message, meta_str)
        elif lvl == "WARNING":
            logger.warning("[RAG][doc=%s] %s %s", document_id, message, meta_str)
        elif lvl in {"ERROR", "CRITICAL"}:
            logger.error("[RAG][doc=%s] %s %s", document_id, message, meta_str)
        else:
            logger.info("[RAG][doc=%s] %s %s", document_id, message, meta_str)

        try:
            with session_scope() as session:
                session.add(
                    RagDocumentLog(
                        document_id=document_id,
                        log_level=level,
                        message=message,
                        meta_info=json.dumps(meta, ensure_ascii=False) if meta else None,
                    )
                )
        except Exception:
            # 日志写入失败不应阻塞主流程
            return

    def prepare_excel_document_upload(
        self,
        collection_id: int,
        filename: str,
        file_path: str,
        file_size: int,
        file_type: str,
        user_id: int,
        overwrite: bool = False,
    ) -> KnowledgeDocument:
        """创建/复用 KnowledgeDocument，并置为 processing。

        真正的 Excel 解析与向量写入由后台任务执行（process_excel_document_by_id）。
        """
        with session_scope() as session:
            collection = session.get(KnowledgeCollection, collection_id)
            if collection is None:
                raise ValueError("知识库不存在")

            collection.chroma_collection_id = collection.chroma_collection_id or f"kc_{collection.id}"
            collection.embedding_model = collection.embedding_model or "bge-base-zh-v1.5"
            session.add(collection)
            session.flush()
            session.refresh(collection)

            existing = (
                session.query(KnowledgeDocument)
                .filter(
                    KnowledgeDocument.collection_id == collection_id,
                    KnowledgeDocument.filename == filename,
                )
                .order_by(KnowledgeDocument.created_at.desc())
                .first()
            )
            if existing is not None and not overwrite:
                raise DocumentAlreadyExistsError("同名文档已存在，确认覆盖后再上传")

            if existing is not None and overwrite:
                # 覆盖：先尝试删除旧向量（失败不影响后续重建）
                try:
                    if collection.chroma_collection_id:
                        self._append_document_log(
                            existing.id,
                            LogLevel.INFO,
                            "检测到覆盖上传：准备删除旧向量索引",
                            {"chroma_collection_id": collection.chroma_collection_id},
                        )
                        self._vector.delete_by_document_id(collection.chroma_collection_id, existing.id)
                        self._append_document_log(existing.id, LogLevel.INFO, "旧向量索引删除完成")
                except Exception:
                    self._append_document_log(existing.id, LogLevel.WARNING, "删除旧向量索引失败（将继续重建）")
                    pass
                document = existing
            else:
                document = KnowledgeDocument(
                    collection_id=collection_id,
                    filename=filename,
                    file_path=file_path,
                    file_size=file_size,
                    file_type=file_type,
                    upload_status="processing",
                    uploaded_by=user_id,
                )
                session.add(document)
                collection.document_count = (collection.document_count or 0) + 1
                session.add(collection)

            # 更新（覆盖/新建都要）
            document.file_path = file_path
            document.file_size = file_size
            document.file_type = file_type
            document.chunk_count = 0
            document.upload_status = "processing"
            document.processing_error = None
            document.meta_info = None
            document.content = (document.content or "")[:0]
            session.add(document)

            session.flush()
            session.refresh(document)

        self._append_document_log(
            document.id,
            LogLevel.INFO,
            "已接收文件，准备后台构建索引",
            {"filename": filename, "file_path": file_path, "file_size": file_size, "overwrite": overwrite},
        )
        return document

    def process_excel_document_by_id(self, document_id: int) -> None:
        """后台任务：根据 document_id 读取文件并构建索引（写入 Chroma）。"""
        start_all = time.time()
        # 1) 读取文档与集合信息
        with session_scope() as session:
            doc = session.get(KnowledgeDocument, document_id)
            if doc is None:
                return
            collection_id = int(doc.collection_id)
            filename = doc.filename
            file_path = doc.file_path
            doc.upload_status = "processing"
            doc.processing_error = None
            session.add(doc)

        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "开始处理任务",
            {"collection_id": collection_id, "filename": filename, "file_path": file_path},
        )

        try:
            t0 = time.time()
            raw_bytes = Path(file_path).read_bytes()
            self._append_document_log(
                document_id,
                LogLevel.INFO,
                "文件读取完成",
                {"bytes": len(raw_bytes), "elapsed_ms": int((time.time() - t0) * 1000)},
            )
        except Exception as exc:
            err = f"读取文件失败：{exc}"
            self._append_document_log(document_id, LogLevel.ERROR, err)
            with session_scope() as session:
                doc = session.get(KnowledgeDocument, document_id)
                if doc:
                    doc.upload_status = "failed"
                    doc.processing_error = err
                    session.add(doc)
            return

        try:
            # 2) 构建向量索引
            self._append_document_log(document_id, LogLevel.INFO, "开始解析 Excel（第一个 Sheet）")
            t_index = time.time()
            report = self._index_excel_to_chroma(
                chroma_collection_id=self._ensure_chroma_collection_id(collection_id),
                document_id=document_id,
                collection_id=collection_id,
                filename=filename,
                raw_bytes=raw_bytes,
            )
            self._append_document_log(
                document_id,
                LogLevel.INFO,
                "Excel 解析 + 向量写入完成",
                {"elapsed_ms": int((time.time() - t_index) * 1000)},
            )
            self._append_document_log(
                document_id,
                LogLevel.INFO,
                "向量索引写入完成，正在更新文档状态",
                {
                    "rows_total": report.rows_total,
                    "rows_indexed": report.rows_indexed,
                    "rows_skipped_empty_tag": report.rows_skipped_empty_tag,
                    "primary_tag_key": report.primary_tag_key,
                    "primary_tag_unique": report.primary_tag_unique,
                },
            )

            # 3) 更新文档记录
            with session_scope() as session:
                doc = session.get(KnowledgeDocument, document_id)
                if doc is None:
                    raise ValueError("文档不存在")
                doc.chunk_count = report.rows_indexed
                doc.upload_status = "completed"
                doc.processing_error = None
                doc.meta_info = json.dumps(
                    {
                        "primary_tag_key": report.primary_tag_key,
                        "sheet_name": report.sheet_name,
                        "rows_total": report.rows_total,
                        "rows_indexed": report.rows_indexed,
                        "rows_skipped_empty_tag": report.rows_skipped_empty_tag,
                        "primary_tag_unique": report.primary_tag_unique,
                    },
                    ensure_ascii=False,
                )
                session.add(doc)

            self._bump_collection_revision(collection_id)
            self._append_document_log(
                document_id,
                LogLevel.INFO,
                "处理完成",
                {"status": "completed", "total_elapsed_ms": int((time.time() - start_all) * 1000)},
            )
        except Exception as exc:
            err = str(exc)
            self._append_document_log(
                document_id,
                LogLevel.ERROR,
                "处理失败",
                {"error": err, "total_elapsed_ms": int((time.time() - start_all) * 1000)},
            )
            with session_scope() as session:
                doc = session.get(KnowledgeDocument, document_id)
                if doc:
                    doc.upload_status = "failed"
                    doc.processing_error = err
                    session.add(doc)

    def get_or_create_default_collection_for_user(self, user_id: int) -> KnowledgeCollection:
        """分析工作流默认用的知识库集合。

        规则：优先用户自己的最新集合，其次公共集合；都没有则创建用户默认集合。
        """
        with session_scope() as session:
            collection = (
                session.query(KnowledgeCollection)
                .filter(KnowledgeCollection.created_by == user_id)
                .order_by(KnowledgeCollection.created_at.desc())
                .first()
            )
            if collection is None:
                collection = (
                    session.query(KnowledgeCollection)
                    .filter(KnowledgeCollection.created_by.is_(None))
                    .order_by(KnowledgeCollection.created_at.desc())
                    .first()
                )
            if collection is None:
                collection = KnowledgeCollection(
                    name="默认知识库",
                    description="系统自动创建",
                    document_count=0,
                    collection_type="document",
                    is_active=True,
                    created_by=user_id,
                )
                session.add(collection)
                session.flush()
            collection.chroma_collection_id = collection.chroma_collection_id or f"kc_{collection.id}"
            collection.embedding_model = collection.embedding_model or "bge-base-zh-v1.5"
            session.add(collection)
            session.refresh(collection)

        self._vector.get_or_create_collection(collection.chroma_collection_id)
        return collection

    def add_excel_document(
        self,
        collection_id: int,
        filename: str,
        file_size: int,
        file_type: str,
        user_id: int,
        raw_bytes: bytes,
        overwrite: bool = False,
    ) -> tuple[KnowledgeDocument, ExcelIndexReport]:
        with session_scope() as session:
            collection = session.get(KnowledgeCollection, collection_id)
            if collection is None:
                raise ValueError("知识库不存在")

            collection.chroma_collection_id = collection.chroma_collection_id or f"kc_{collection.id}"
            collection.embedding_model = collection.embedding_model or "bge-base-zh-v1.5"
            session.add(collection)

            existing = (
                session.query(KnowledgeDocument)
                .filter(
                    KnowledgeDocument.collection_id == collection_id,
                    KnowledgeDocument.filename == filename,
                )
                .order_by(KnowledgeDocument.created_at.desc())
                .first()
            )
            if existing is not None and not overwrite:
                raise DocumentAlreadyExistsError("同名文档已存在，确认覆盖后再上传")

            # 覆盖：删除旧向量 chunk
            if existing is not None and overwrite:
                if collection.chroma_collection_id:
                    self._vector.delete_by_document_id(collection.chroma_collection_id, existing.id)
                document = existing
            else:
                document = KnowledgeDocument(
                    collection_id=collection_id,
                    filename=filename,
                    file_path=f"/knowledge-docs/{collection_id}/{filename}",
                    file_size=file_size,
                    file_type=file_type,
                    upload_status="processing",
                    uploaded_by=user_id,
                )
                session.add(document)
                # 新增才增加计数
                collection.document_count = (collection.document_count or 0) + 1
                session.add(collection)

            # 先拿到 document.id，便于 chunk id 与 metadata 绑定
            session.flush()
            session.refresh(document)

        # 解析 Excel + 写入 Chroma（放到事务外，避免长事务与模型编码占用）
        report = self._index_excel_to_chroma(
            chroma_collection_id=self._ensure_chroma_collection_id(collection_id),
            document_id=document.id,
            collection_id=collection_id,
            filename=filename,
            raw_bytes=raw_bytes,
        )

        # 入库完成后更新 document 状态与摘要，并 bump revision 让缓存自然失效
        with session_scope() as session:
            doc = session.get(KnowledgeDocument, document.id)
            if doc is None:
                raise ValueError("文档不存在")
            doc.file_size = file_size
            doc.file_type = file_type
            doc.chunk_count = report.rows_indexed
            # content 仅用于后台审计/快速预览，不作为检索来源
            doc.content = (doc.content or "")[:0]
            doc.upload_status = "completed"
            doc.processing_error = None
            doc.meta_info = json.dumps(
                {
                    "primary_tag_key": report.primary_tag_key,
                    "sheet_name": report.sheet_name,
                    "rows_total": report.rows_total,
                    "rows_indexed": report.rows_indexed,
                    "rows_skipped_empty_tag": report.rows_skipped_empty_tag,
                    "primary_tag_unique": report.primary_tag_unique,
                    "overwritten": bool(overwrite and existing is not None),
                },
                ensure_ascii=False,
            )
            session.add(doc)
            session.flush()
            session.refresh(doc)
            final_doc = doc

        self._bump_collection_revision(collection_id)
        return final_doc, report

    def _ensure_chroma_collection_id(self, collection_id: int) -> str:
        with session_scope() as session:
            collection = session.get(KnowledgeCollection, collection_id)
            if collection is None:
                raise ValueError("知识库不存在")
            collection.chroma_collection_id = collection.chroma_collection_id or f"kc_{collection.id}"
            session.add(collection)
            session.flush()
            session.refresh(collection)
            chroma_id = collection.chroma_collection_id
        self._vector.get_or_create_collection(chroma_id)
        return chroma_id

    def _collection_revision_key(self, collection_id: int) -> str:
        return f"rag:rev:{collection_id}"

    def _get_collection_revision(self, collection_id: int) -> int:
        key = self._collection_revision_key(collection_id)
        cached = self._cache.get_json(key)
        if cached and isinstance(cached.get("value"), int):
            return int(cached["value"])
        # 不存在时初始化为 0（不写入也可）
        return 0

    def _bump_collection_revision(self, collection_id: int) -> int:
        key = self._collection_revision_key(collection_id)
        val = self._cache.incr(key)
        # incr 后确保以 JSON 结构存一份，便于 MemoryTTLCache 一致读取
        self._cache.set_json(key, {"value": int(val)}, ttl_seconds=24 * 3600)
        return int(val)

    def _index_excel_to_chroma(
        self,
        chroma_collection_id: str,
        document_id: int,
        collection_id: int,
        filename: str,
        raw_bytes: bytes,
    ) -> ExcelIndexReport:
        t0 = time.time()
        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "开始加载 Excel 工作簿",
            {"bytes": len(raw_bytes), "chroma_collection_id": chroma_collection_id},
        )
        wb = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        sheet_name = ws.title
        self._append_document_log(document_id, LogLevel.INFO, "已打开 Excel（第一个 Sheet）", {"sheet_name": sheet_name})

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            raise ValueError("Excel 第一行为空，无法识别表头")

        headers = [normalize_text(h) for h in header_row]
        if not headers or not headers[0]:
            headers = [f"列{i+1}" for i in range(len(header_row))]
        primary_tag_key = headers[0] or "主标签"
        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "表头解析完成",
            {"primary_tag_key": primary_tag_key, "header_count": len(headers)},
        )

        ids: list[str] = []
        docs: list[str] = []
        metas: list[dict[str, Any]] = []

        rows_total = 0
        rows_indexed = 0
        rows_skipped_empty_tag = 0
        primary_values: set[str] = set()

        # 从第 2 行开始，按行写入
        for idx, row in enumerate(rows_iter, start=2):
            rows_total += 1
            # 对齐列数
            row_values = list(row) + [None] * max(0, len(headers) - len(row))
            primary_raw = row_values[0] if row_values else None
            primary_val = normalize_primary_tag_value(primary_raw)
            if not primary_val:
                rows_skipped_empty_tag += 1
                continue

            primary_values.add(primary_val)
            kv_pairs = []
            for col_idx, col_name in enumerate(headers):
                if not col_name:
                    continue
                cell = row_values[col_idx] if col_idx < len(row_values) else None
                if cell is None:
                    continue
                cell_text = normalize_text(cell)
                if not cell_text:
                    continue
                kv_pairs.append(f"{col_name}={cell_text}")

            # 文本构建：主标签置顶，其他列按 key=value 展开
            row_text = f"【{primary_tag_key}={primary_val}】" + ("；" + "；".join(kv_pairs) if kv_pairs else "")

            chunk_id = f"doc_{document_id}_row_{idx}"
            ids.append(chunk_id)
            docs.append(row_text)
            metas.append(
                {
                    "collection_id": int(collection_id),
                    "document_id": int(document_id),
                    "source_filename": filename,
                    "sheet_name": sheet_name,
                    "row_index": int(idx),
                    "primary_tag_key": primary_tag_key,
                    "primary_tag_value": primary_val,
                }
            )
            rows_indexed += 1

        if rows_indexed == 0:
            raise ValueError("未索引任何有效行（可能第一列全为空）")

        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "Excel 行读取完成，准备 embedding",
            {
                "rows_total": rows_total,
                "rows_indexed": rows_indexed,
                "rows_skipped_empty_tag": rows_skipped_empty_tag,
                "primary_tag_unique": len(primary_values),
            },
        )

        # --- embedding（分批 + 进度日志）---
        total_rows = len(docs)
        self._append_document_log(document_id, LogLevel.INFO, "开始 embedding", {"rows": total_rows, "batch_size": 32})
        t_embed = time.time()
        last_percent = -1

        def _on_embed_batch(done: int, total: int, batch_ms: int) -> None:
            nonlocal last_percent
            percent = int(done * 100 / max(1, total))
            # 限流：每提升 5% 或最后一次才记录
            if percent >= last_percent + 5 or done >= total:
                last_percent = percent
                self._append_document_log(
                    document_id,
                    LogLevel.INFO,
                    f"embedding 进度 {done}/{total}（{percent}%）",
                    {"done": done, "total": total, "percent": percent, "batch_ms": batch_ms},
                )

        embeddings = embed_texts_batched(docs, batch_size=32, on_batch=_on_embed_batch)
        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "embedding 完成，准备写入向量库",
            {"rows": total_rows, "elapsed_ms": int((time.time() - t_embed) * 1000)},
        )

        # --- 写入向量库（分批 + 进度日志）---
        # Windows/小型磁盘环境下，单次 upsert 过大可能导致长时间阻塞（看起来“没进度”）
        upsert_batch = 128
        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "开始写入向量库",
            {"rows": total_rows, "batch_size": upsert_batch},
        )
        t_upsert = time.time()
        upserted = 0
        total_batches = (total_rows + upsert_batch - 1) // upsert_batch
        for batch_idx, start in enumerate(range(0, total_rows, upsert_batch), start=1):
            end = min(total_rows, start + upsert_batch)
            percent_before = int(start * 100 / max(1, total_rows))
            self._append_document_log(
                document_id,
                LogLevel.INFO,
                f"写入向量库开始（batch {batch_idx}/{total_batches}，{start}/{total_rows}，{percent_before}%）",
                {"batch": batch_idx, "total_batches": total_batches, "start": start, "end": end},
            )
            t_batch = time.time()
            try:
                self._vector.upsert(
                    chroma_collection_id=chroma_collection_id,
                    ids=ids[start:end],
                    documents=docs[start:end],
                    metadatas=metas[start:end],
                    embeddings=embeddings[start:end],
                )
            except Exception as exc:
                self._append_document_log(
                    document_id,
                    LogLevel.ERROR,
                    f"写入向量库失败（batch {batch_idx}/{total_batches}）",
                    {"error": str(exc), "batch": batch_idx, "start": start, "end": end},
                )
                raise

            upserted = end
            percent = int(upserted * 100 / max(1, total_rows))
            self._append_document_log(
                document_id,
                LogLevel.INFO,
                f"写入向量库完成（batch {batch_idx}/{total_batches}，{upserted}/{total_rows}，{percent}%）",
                {"done": upserted, "total": total_rows, "percent": percent, "batch_ms": int((time.time() - t_batch) * 1000)},
            )

        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "向量库写入完成",
            {"rows": total_rows, "elapsed_ms": int((time.time() - t_upsert) * 1000)},
        )
        try:
            wb.close()
        except Exception:
            pass

        self._append_document_log(
            document_id,
            LogLevel.INFO,
            "索引构建阶段结束",
            {"elapsed_ms": int((time.time() - t0) * 1000)},
        )

        return ExcelIndexReport(
            primary_tag_key=primary_tag_key,
            sheet_name=sheet_name,
            rows_total=rows_total,
            rows_indexed=rows_indexed,
            rows_skipped_empty_tag=rows_skipped_empty_tag,
            primary_tag_unique=len(primary_values),
        )

    def list_documents(self, collection_id: int) -> Sequence[KnowledgeDocument]:
        with session_scope() as session:
            return (
                session.query(KnowledgeDocument)
                .filter(KnowledgeDocument.collection_id == collection_id)
                .order_by(KnowledgeDocument.created_at.desc())
                .all()
            )

    def list_queries(self, collection_id: int, limit: int = 50) -> Sequence[RAGQuery]:
        with session_scope() as session:
            return (
                session.query(RAGQuery)
                .filter(RAGQuery.collection_id == collection_id)
                .order_by(RAGQuery.created_at.desc())
                .limit(limit)
                .all()
            )

    def query(
        self,
        collection_id: int,
        query: str,
        user_id: int | None = None,
        limit: int = 5,
        show_retrieval: bool = True,
    ) -> dict:
        start = time.time()

        chroma_id = self._ensure_chroma_collection_id(collection_id)
        revision = self._get_collection_revision(collection_id)
        normalized_query = normalize_text(query)
        cache_key = f"rag:q:{collection_id}:r{revision}:k{limit}:{normalized_query}"

        cached = self._cache.get_json(cache_key)
        if cached:
            # 仍然写入 query history（缓存命中也要记录耗时/结果数）
            elapsed_ms = int((time.time() - start) * 1000)
            self._record_query(collection_id, query, cached.get("documents", []), cached.get("response", ""), user_id, elapsed_ms)
            return {
                "response": cached.get("response", ""),
                "documents": (cached.get("documents") or []) if show_retrieval else [],
                "query_time": elapsed_ms,
            }

        # 1) 严格等值命中优先（避免跨表错配）：
        #    - 若 query 形如 "停充码ChrgEndNum 8bit=10"，提取 (key,val)，并同时约束 primary_tag_key + primary_tag_value
        #    - 若 query 仅为纯数字（例如用户直接输入 "10"），允许按 value-only 严格命中
        normalized_query_for_kv = normalize_text(query)
        candidate_kv = extract_candidate_primary_tag_kv(normalized_query_for_kv)
        candidate_tag: str | None = None
        candidate_key: str | None = None
        if candidate_kv:
            candidate_key, candidate_tag = candidate_kv[0], candidate_kv[1]
        else:
            # 仅纯数字场景才启用 value-only strict（否则容易误把任意 "=2" 命中成停充码2）
            if normalized_query_for_kv.isdigit():
                candidate_tag = normalize_primary_tag_value(normalized_query_for_kv)
        results: list[dict[str, Any]] = []
        if candidate_tag:
            strict_cache_key = (
                f"rag:strict:{collection_id}:r{revision}:{candidate_key}={candidate_tag}"
                if candidate_key
                else f"rag:strict:{collection_id}:r{revision}:{candidate_tag}"
            )
            strict_cached = self._cache.get_json(strict_cache_key)
            if strict_cached and isinstance(strict_cached.get("documents"), list):
                results = strict_cached["documents"]
            else:
                # Chroma where 语法：顶层只能是一个操作符。
                # 需要同时约束 key + value 时，使用 $and。
                where: dict[str, Any]
                if candidate_key:
                    where = {"$and": [{"primary_tag_key": candidate_key}, {"primary_tag_value": candidate_tag}]}
                else:
                    where = {"primary_tag_value": candidate_tag}
                strict_hits = self._vector.get_by_where(chroma_collection_id=chroma_id, where=where, limit=50)
                results = self._format_hits(strict_hits, prefer_strict_score=True)[:limit]
                # 负缓存：严格等值未命中也缓存短 TTL，避免击穿
                self._cache.set_json(strict_cache_key, {"documents": results}, ttl_seconds=60 if not results else 3600)

        # 2) 若无严格命中，走向量检索
        if not results:
            emb = embed_query(normalized_query)
            hits = self._vector.query(chroma_collection_id=chroma_id, query_embedding=emb, top_k=limit)
            results = self._format_hits(hits, prefer_strict_score=False)

        response_text = self._build_rule_answer(query=normalized_query, documents=results, candidate_tag=candidate_tag)
        elapsed_ms = int((time.time() - start) * 1000)

        payload = {
            "response": response_text,
            # 缓存中始终保留证据链，show_retrieval 只影响返回给客户端的展示
            "documents": results,
            "query_time": elapsed_ms,
        }
        self._cache.set_json(cache_key, payload, ttl_seconds=10 * 60)
        self._record_query(collection_id, query, results, response_text, user_id, elapsed_ms)
        return {
            "response": payload["response"],
            "documents": payload["documents"] if show_retrieval else [],
            "query_time": payload["query_time"],
        }

    def _format_hits(self, hits: Sequence[VectorHit], prefer_strict_score: bool) -> list[dict[str, Any]]:
        formatted: list[dict[str, Any]] = []
        for hit in hits:
            meta = hit.metadata or {}
            distance = hit.distance
            if prefer_strict_score:
                score = 1.0
            else:
                # 使用单调映射到 [0,1]，避免不同 distance 空间导致“分数不可解释”
                score = None
                if distance is not None:
                    score = float(1.0 / (1.0 + max(0.0, distance)))
            content = hit.document or ""
            formatted.append(
                {
                    "id": hit.id,
                    "content": content,
                    "snippet": content[:500],
                    "score": score if score is not None else 0.0,
                    "distance": distance,
                    "metadata": meta,
                    "filename": meta.get("source_filename"),
                    "sheet_name": meta.get("sheet_name"),
                    "row_index": meta.get("row_index"),
                    "primary_tag_key": meta.get("primary_tag_key"),
                    "primary_tag_value": meta.get("primary_tag_value"),
                }
            )
        # 若严格等值返回多行，按信息完整度/行号做排序
        formatted.sort(
            key=lambda x: (
                -(len(str(x.get("content") or ""))),
                int(x.get("row_index") or 10**9),
            )
        )
        return formatted

    def _build_rule_answer(self, query: str, documents: list[dict[str, Any]], candidate_tag: str | None) -> str:
        if not documents:
            return "未找到相关条目（知识库为空或不匹配）。"
        top = documents[0]
        key = top.get("primary_tag_key") or "主标签"
        val = top.get("primary_tag_value") or (candidate_tag or "")
        snippet = top.get("snippet") or ""
        # 给出“规则化答案”：优先展示主标签与来源
        src = top.get("filename") or "未知文件"
        row = top.get("row_index")
        row_part = f"（第{row}行）" if row else ""
        return f"{key}={val}{row_part}：{snippet}（来源：{src}）"

    def _record_query(
        self,
        collection_id: int,
        query: str,
        documents: list[dict[str, Any]],
        response_text: str,
        user_id: int | None,
        elapsed_ms: int,
    ) -> None:
        try:
            with session_scope() as session:
                session.add(
                    RAGQuery(
                        collection_id=collection_id,
                        query_text=query,
                        result_count=len(documents or []),
                        response_text=response_text,
                        user_id=user_id,
                        query_time_ms=elapsed_ms,
                    )
                )
        except Exception:
            # 查询日志失败不应影响主流程
            return


_rag_service: RagService | None = None


def get_rag_service() -> RagService:
    global _rag_service
    if _rag_service is None:
        _rag_service = RagService()
    return _rag_service
